"""
Excel 数据处理服务
基于 pandas + openpyxl + DuckDB，支持异步加载、Schema提取、安全脚本执行和多轮SQL查询

升级特性:
  1. 智能表头识别 — 评分机制自动识别表头行（借鉴 DB-GPT ExcelKnowledge._find_header_row）
  2. 合并单元格处理 — 自动展开 merged_cells.ranges
  3. DuckDB 查询引擎 — 将 Excel 数据导入 DuckDB 内存数据库，支持自然语言→SQL→多轮交互查询
"""

import asyncio
import base64
import contextlib
import io
import logging
import os
import traceback
from typing import Any, Dict, List, Optional, Tuple

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import pandas as pd

logger = logging.getLogger(__name__)

_SAFE_BUILTINS = {
    "abs": abs,
    "all": all,
    "any": any,
    "bool": bool,
    "dict": dict,
    "enumerate": enumerate,
    "filter": filter,
    "float": float,
    "int": int,
    "isinstance": isinstance,
    "len": len,
    "list": list,
    "map": map,
    "max": max,
    "min": min,
    "print": print,
    "range": range,
    "round": round,
    "set": set,
    "sorted": sorted,
    "str": str,
    "sum": sum,
    "tuple": tuple,
    "type": type,
    "zip": zip,
}

_FORBIDDEN_NAMES = {
    "eval", "compile", "exec",
    "breakpoint", "exit", "quit",
}

_FORBIDDEN_MODULES = {
    "os", "sys", "subprocess", "shutil", "pathlib",
    "socket", "http", "urllib", "requests",
    "pickle", "shelve", "marshal", "ctypes",
}


class ExcelDataProcessor:
    """Excel 异步数据处理器"""

    async def load_excel_to_dataframe(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        header: int = 0,
    ) -> Dict[str, Any]:
        result = await asyncio.to_thread(
            self._load_sync, file_path, sheet_name, header
        )
        logger.info(
            f"Excel loaded: {file_path}, shape={result['shape']}, "
            f"sheets={result['sheet_names']}"
        )
        return result

    async def extract_excel_schema(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        sample_rows: int = 5,
    ) -> Dict[str, Any]:
        """
        提取 Excel 的 Schema 信息，返回 Markdown 和 JSON 格式

        Args:
            file_path: Excel 文件路径
            sheet_name: 工作表名
            sample_rows: 采样行数

        Returns:
            {
                "schema_markdown": str,   # Markdown 格式的 Schema
                "schema_json": dict,      # JSON 结构化 Schema
                "columns": List[str],
                "dtypes": Dict[str, str],
                "null_counts": Dict[str, int],
                "sample_data": List[Dict],
                "shape": tuple,
                "sheet_names": List[str],
                "unique_counts": Dict[str, int],
                "stats": Dict[str, Any],
            }
        """
        result = await asyncio.to_thread(
            self._extract_schema_sync, file_path, sheet_name, sample_rows
        )
        logger.info(
            f"Schema extracted: {file_path}, columns={len(result['columns'])}, "
            f"shape={result['shape']}"
        )
        return result

    async def execute_pandas_code(
        self,
        code: str,
        input_path: str,
        output_path: str,
    ) -> Dict[str, Any]:
        """
        在受限沙盒中执行 LLM 生成的 Pandas 代码

        代码中可使用:
          - df: 已加载的 DataFrame (从 input_path 读取)
          - pd: pandas 库
          - np: numpy 库
          - input_path: 输入文件路径 (str)
          - output_path: 输出文件路径 (str)

        代码需将结果 DataFrame 保存到 output_path

        Args:
            code: LLM 生成的 Python 代码
            input_path: 输入 Excel 路径
            output_path: 输出 Excel 路径

        Returns:
            {
                "success": bool,
                "output_path": Optional[str],
                "error": Optional[str],
                "log": List[str],
                "original_shape": tuple,
                "result_shape": Optional[tuple],
            }
        """
        return await asyncio.to_thread(
            self._execute_pandas_sync, code, input_path, output_path
        )

    async def execute_cleaning_script(
        self,
        df: pd.DataFrame,
        pandas_code_string: str,
    ) -> Dict[str, Any]:
        logs: List[str] = []

        def _print_capture(*args, **kwargs):
            logs.append(" ".join(str(a) for a in args))

        sandbox_globals = {
            "__builtins__": __builtins__,
            **_SAFE_PANDAS,
            "df": df.copy(),
            "np": __import__("numpy"),
            "print": _print_capture,
        }

        try:
            exec(pandas_code_string, sandbox_globals)
            result_df = sandbox_globals.get("df", df)

            if not isinstance(result_df, pd.DataFrame):
                return {
                    "dataframe": df,
                    "shape": df.shape,
                    "success": False,
                    "error": "代码执行后 df 不是 DataFrame 类型",
                    "log": logs,
                }

            logger.info(f"Cleaning script executed: {df.shape} -> {result_df.shape}")
            return {
                "dataframe": result_df,
                "shape": result_df.shape,
                "success": True,
                "error": None,
                "log": logs,
            }

        except Exception as e:
            logger.error(f"Cleaning script failed: {e}")
            return {
                "dataframe": df,
                "shape": df.shape,
                "success": False,
                "error": str(e),
                "log": logs,
            }

    async def execute_python_code(self, file_path: str, code: str, user_instruction: str = "") -> Dict[str, Any]:
        return await asyncio.to_thread(self._execute_python_sync, file_path, code, user_instruction)

    # ================================================================
    # Text-to-SQL-to-Chart 管线：DuckDB 查询 + Python 系统级模板绘图
    # ================================================================

    async def execute_chart_pipeline(
        self,
        file_path: str,
        sql: str,
        chart_type: str = "",
        title: str = "数据分析",
        text_analysis: str = "",
    ) -> Dict[str, Any]:
        """执行 Text-to-SQL-to-Chart 管线。

        1. 加载 Excel → DuckDB 内存表
        2. 执行 LLM 生成的 SQL 获取聚合结果
        3. Python 系统代码根据 chart_type 模板绘图
        4. 返回 {success, output, image_base64}

        Args:
            file_path: Excel 文件路径
            sql: DuckDB SQL 查询（表名固定为 df）
            chart_type: pie/bar/line/scatter/hist/box 或空字符串（不画图）
            title: 图表标题
            text_analysis: LLM 的文字分析
        """
        return await asyncio.to_thread(
            self._execute_chart_pipeline_sync,
            file_path, sql, chart_type, title, text_analysis,
        )

    def _execute_chart_pipeline_sync(
        self,
        file_path: str,
        sql: str,
        chart_type: str,
        title: str,
        text_analysis: str,
    ) -> Dict[str, Any]:
        import duckdb
        import numpy as np

        # --- 1. 加载 DataFrame ---
        try:
            df = self._load_dataframe(file_path)
        except Exception as e:
            return {"success": False, "error": f"Failed to read file: {e}", "output": "", "image_base64": None}

        # --- 2. DuckDB SQL 查询 ---
        try:
            conn = duckdb.connect(":memory:")
            conn.register("df", df)
            result_df = conn.execute(sql).fetchdf()
            conn.close()
            logger.info(f"📊 [ChartPipeline] SQL 执行成功, 结果 {result_df.shape[0]} 行 × {result_df.shape[1]} 列")
        except Exception as e:
            return {"success": False, "error": f"SQL 执行失败: {e}\nSQL: {sql}", "output": "", "image_base64": None}

        # --- 3. 将查询结果转为 Markdown（供 Step3 报告生成使用） ---
        try:
            result_markdown = result_df.to_markdown(index=False)
        except Exception:
            result_markdown = result_df.to_string(index=False, max_rows=30)

        # --- 4. 构建文字输出（仅使用 LLM 的深度分析，不再追加原始 SQL 结果） ---
        output = text_analysis if text_analysis else "（未生成分析报告）"

        # --- 4. 系统级模板绘图 ---
        img_base64 = None
        if chart_type and chart_type != "none":
            try:
                # 配置中文字体
                plt.rcParams['font.sans-serif'] = ['SimHei', 'Noto Sans CJK JP', 'Droid Sans Fallback', 'DejaVu Sans', 'sans-serif']
                plt.rcParams['axes.unicode_minus'] = False
                plt.clf()

                ncols = result_df.shape[1]
                nrows = result_df.shape[0]

                # 空结果保护
                if nrows == 0:
                    logger.warning(f"📊 [ChartPipeline] SQL 查询结果为空（0行），跳过绘图")
                    return {"success": True, "output": output, "image_base64": None, "result_markdown": result_markdown}

                # 多列警告：SQL 返回超过2列时，只取前2列绘图
                if ncols > 2:
                    logger.warning(f"📊 [ChartPipeline] SQL 返回 {ncols} 列，取前2列绘图（第1列=标签，第2列=数值）")

                # 统一提取前两列（iloc 绝对安全，不依赖列名）
                col0 = result_df.iloc[:, 0]  # 第1列：分类标签 / X轴
                col1 = result_df.iloc[:, 1] if ncols >= 2 else None  # 第2列：数值 / Y轴

                if chart_type == "pie":
                    if col1 is not None:
                        labels = col0.astype(str).values
                        values = col1.values
                    else:
                        # 单列：自动计数
                        counts = col0.value_counts()
                        labels = counts.index.astype(str).values
                        values = counts.values
                    plt.figure(figsize=(8, 8))
                    plt.pie(values, labels=labels, autopct='%1.1f%%', startangle=90,
                            colors=plt.cm.Set3.colors[:len(labels)])
                    plt.title(title, fontsize=14)

                elif chart_type == "bar":
                    if col1 is not None:
                        labels = col0.astype(str).values
                        values = col1.values
                    else:
                        counts = col0.value_counts()
                        labels = counts.index.astype(str).values
                        values = counts.values
                    plt.figure(figsize=(10, 6))
                    bars = plt.bar(range(len(labels)), values, color=plt.cm.Paired.colors[:len(labels)],
                                   edgecolor='white', linewidth=0.5)
                    plt.xticks(range(len(labels)), labels, rotation=45, ha='right')
                    plt.title(title, fontsize=14)
                    plt.ylabel(result_df.columns[1] if ncols >= 2 else 'Count', fontsize=12)
                    plt.grid(True, alpha=0.3, axis='y')
                    # 在柱子上方显示数值
                    for bar_item, val in zip(bars, values):
                        plt.text(bar_item.get_x() + bar_item.get_width() / 2, bar_item.get_height(),
                                 f'{val}', ha='center', va='bottom', fontsize=9)

                elif chart_type == "line":
                    if col1 is not None:
                        x_vals = col0.values
                        y_vals = col1.values
                    else:
                        x_vals = range(len(result_df))
                        y_vals = col0.values
                    plt.figure(figsize=(10, 6))
                    plt.plot(x_vals, y_vals, marker='o', linewidth=2, color='steelblue', markersize=6)
                    plt.fill_between(range(len(y_vals)), y_vals, alpha=0.1, color='steelblue')
                    plt.title(title, fontsize=14)
                    plt.xlabel(result_df.columns[0] if ncols >= 1 else '', fontsize=12)
                    plt.ylabel(result_df.columns[1] if ncols >= 2 else '', fontsize=12)
                    plt.grid(True, alpha=0.3)
                    plt.xticks(rotation=45, ha='right')

                elif chart_type == "scatter":
                    if col1 is not None:
                        x_vals = col0.values
                        y_vals = col1.values
                    else:
                        # 单列散点图：用行号做 X 轴
                        x_vals = range(len(result_df))
                        y_vals = col0.values
                    plt.figure(figsize=(10, 6))
                    plt.scatter(x_vals, y_vals, alpha=0.6, c='steelblue', edgecolors='navy', s=50)
                    plt.title(title, fontsize=14)
                    plt.xlabel(result_df.columns[0] if ncols >= 1 else 'Index', fontsize=12)
                    plt.ylabel(result_df.columns[1] if ncols >= 2 else result_df.columns[0], fontsize=12)
                    plt.grid(True, alpha=0.3)

                elif chart_type == "hist":
                    # 直方图：取第1列数值
                    values = col0.dropna().values
                    plt.figure(figsize=(10, 6))
                    n, bins, patches = plt.hist(values, bins=20, color='steelblue', edgecolor='white', alpha=0.7)
                    plt.title(title, fontsize=14)
                    plt.xlabel(result_df.columns[0], fontsize=12)
                    plt.ylabel('频次', fontsize=12)
                    plt.grid(True, alpha=0.3, axis='y')

                elif chart_type == "box":
                    # 箱线图：取第1列数值
                    values = col0.dropna().values
                    plt.figure(figsize=(10, 6))
                    plt.boxplot(values, patch_artist=True,
                                boxprops=dict(facecolor='lightsteelblue', color='navy'),
                                medianprops=dict(color='red', linewidth=2),
                                whiskerprops=dict(color='navy'),
                                capprops=dict(color='navy'))
                    plt.title(title, fontsize=14)
                    plt.ylabel(result_df.columns[0], fontsize=12)
                    plt.grid(True, alpha=0.3, axis='y')

                else:
                    logger.warning(f"📊 [ChartPipeline] 不支持的图表类型: {chart_type}，跳过绘图")
                    return {"success": True, "output": output, "image_base64": None, "result_markdown": result_markdown}

                # --- 5. 保存为 base64 ---
                img_buf = io.BytesIO()
                plt.savefig(img_buf, format='png', bbox_inches='tight', dpi=150)
                img_buf.seek(0)
                img_base64 = base64.b64encode(img_buf.read()).decode('utf-8')
                logger.info(f"📊 [ChartPipeline] 图表生成成功 (base64 len={len(img_base64)})")

            except Exception as e:
                logger.error(f"📊 [ChartPipeline] 绘图失败: {e}", exc_info=True)
                # 绘图失败不影响文字分析结果
                img_base64 = None
            finally:
                try:
                    plt.close('all')
                except Exception:
                    pass

        return {"success": True, "output": output, "image_base64": img_base64, "result_markdown": result_markdown}

    async def get_sheet_names(self, file_path: str) -> List[str]:
        return await asyncio.to_thread(self._get_sheets_sync, file_path)

    async def get_summary(self, df: pd.DataFrame) -> Dict[str, Any]:
        return {
            "shape": df.shape,
            "columns": df.columns.tolist(),
            "dtypes": {col: str(dt) for col, dt in df.dtypes.items()},
            "null_counts": df.isnull().sum().to_dict(),
            "numeric_summary": df.describe().to_dict() if df.select_dtypes(include="number").shape[1] > 0 else {},
        }

    def _load_dataframe(self, file_path: str):
        if file_path.lower().endswith('.csv'):
            try:
                return pd.read_csv(file_path)
            except UnicodeDecodeError:
                return pd.read_csv(file_path, encoding='gbk')
        else:
            return pd.read_excel(file_path)

    # ──────────────────────────────────────────────────────────────
    # 智能表头识别 + 合并单元格处理（借鉴 DB-GPT ExcelKnowledge）
    # ──────────────────────────────────────────────────────────────

    def _find_header_row(
        self, data: List[List[Any]], max_rows_to_check: int = 5
    ) -> Tuple[List[str], int]:
        """
        智能识别表头行 — 评分机制

        对前 max_rows_to_check 行评分:
          - 非数字列越多得分越高 (×2)
          - 唯一值越多得分越高
          - 过滤掉数字为主的行、平均长度过短的行

        Returns:
            (headers, header_row_index)  header_row_index=-1 表示未找到
        """
        if not data or not data[0]:
            return [], -1

        num_cols = len(data[0])
        best_headers: List[str] = []
        best_row_index = -1
        highest_score = -1

        for r_idx in range(min(len(data), max_rows_to_check)):
            row_data = data[r_idx]

            # 跳过空行过多的行
            non_empty = sum(1 for x in row_data if x is not None and str(x).strip() != "")
            if non_empty < num_cols / 4:
                continue

            potential_headers = [
                str(x).strip() if x is not None else "" for x in row_data
            ]

            # 清洗：过滤掉纯数字的"表头"
            cleaned_headers = [
                h for h in potential_headers
                if h and not h.replace(".", "", 1).isdigit()
            ]

            if not any(h for h in cleaned_headers):
                continue

            # 评分：非数字列数 × 2 + 唯一值数
            score = len(cleaned_headers) * 2 + len(set(cleaned_headers))

            # 过滤平均长度过短且列数不足的行
            avg_len = sum(len(h) for h in cleaned_headers) / (
                len(cleaned_headers) if cleaned_headers else 1
            )
            if avg_len < 2 and len(cleaned_headers) < num_cols / 2:
                continue

            if score > highest_score:
                highest_score = score
                best_headers = potential_headers
                best_row_index = r_idx

        if not best_headers:
            logger.warning(
                f"No clear header row found in the first {max_rows_to_check} rows. "
                f"Using default numeric columns."
            )
            best_headers = [str(i) for i in range(num_cols)]
            best_row_index = -1

        # 处理重复列名（自动加序号）
        final_headers: List[str] = []
        header_counts: Dict[str, int] = {}
        for h in best_headers:
            original_h = h
            if h in header_counts:
                header_counts[h] += 1
                h = f"{original_h} ({header_counts[original_h]})"
            else:
                header_counts[h] = 0
            final_headers.append(h if h else f"Col_{len(final_headers)}")

        return final_headers, best_row_index

    def _unmerge_cells(self, sheet) -> List[List[Any]]:
        """
        读取工作表并展开合并单元格

        合并单元格的值会被填充到合并范围内的所有单元格
        """
        max_row = sheet.max_row
        max_col = sheet.max_column

        if max_row == 0 or max_col == 0:
            return []

        unmerged_data = [[None for _ in range(max_col)] for _ in range(max_row)]

        for r_idx, row_cells in enumerate(
            sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)
        ):
            for c_idx, cell in enumerate(row_cells):
                unmerged_data[r_idx][c_idx] = cell.value

        # 展开合并单元格
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col_merged, max_row_merged = merged_range.bounds
            merged_value = sheet.cell(row=min_row, column=min_col).value

            for r in range(min_row, max_row_merged + 1):
                for c in range(min_col, max_col_merged + 1):
                    if r - 1 < len(unmerged_data) and c - 1 < len(unmerged_data[0]):
                        unmerged_data[r - 1][c - 1] = merged_value

        return unmerged_data

    def load_excel_smart(
        self, file_path: str, sheet_name: Optional[str] = None
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        智能加载 Excel — 自动识别表头行 + 处理合并单元格

        Returns:
            (dataframe, metadata)
            metadata 包含: sheet_names, header_row, detected_headers, original_shape
        """
        if file_path.lower().endswith('.csv'):
            df = self._load_dataframe(file_path)
            return df, {
                "sheet_names": ["Sheet1"],
                "header_row": 0,
                "detected_headers": df.columns.tolist(),
                "original_shape": df.shape,
                "merged_cells_count": 0,
            }

        import openpyxl
        workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        sheet_names = workbook.sheetnames
        target = sheet_name or sheet_names[0]
        sheet = workbook[target]

        # 展开合并单元格
        unmerged_data = self._unmerge_cells(sheet)
        merged_count = len(sheet.merged_cells.ranges)

        if not unmerged_data:
            return pd.DataFrame(), {
                "sheet_names": sheet_names,
                "header_row": -1,
                "detected_headers": [],
                "original_shape": (0, 0),
                "merged_cells_count": merged_count,
            }

        # 智能识别表头行
        headers, header_row_idx = self._find_header_row(unmerged_data)

        if header_row_idx != -1:
            data_rows = unmerged_data[header_row_idx + 1:]
            logger.info(
                f"Smart load: sheet '{target}', header at row {header_row_idx + 1}, "
                f"{len(data_rows)} data rows, {merged_count} merged cells expanded"
            )
        else:
            data_rows = unmerged_data
            logger.info(f"Smart load: sheet '{target}', no clear header found")

        # 对齐列数
        num_cols = len(unmerged_data[0]) if unmerged_data else 0
        processed_data_rows = []
        for row in data_rows:
            if len(row) < num_cols:
                processed_data_rows.append(row + [None] * (num_cols - len(row)))
            elif len(row) > num_cols:
                processed_data_rows.append(row[:num_cols])
            else:
                processed_data_rows.append(row)

        df = pd.DataFrame(processed_data_rows)
        if headers:
            df.columns = headers

        # 清理全空列和 Unnamed 列
        df.dropna(axis=1, how="all", inplace=True)
        df = df.loc[:, ~df.columns.str.contains("^Unnamed:", na=False, regex=True)]

        workbook.close()

        return df, {
            "sheet_names": sheet_names,
            "header_row": header_row_idx,
            "detected_headers": df.columns.tolist(),
            "original_shape": df.shape,
            "merged_cells_count": merged_count,
        }

    # ──────────────────────────────────────────────────────────────
    # DuckDB 查询引擎 — ChatExcel 多轮对话查询
    # ──────────────────────────────────────────────────────────────

    def load_to_duckdb(
        self,
        file_path: str,
        table_name: str = "excel_data",
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        将 Excel 数据导入 DuckDB 内存数据库

        使用智能加载（表头识别 + 合并单元格处理）后导入 DuckDB

        Returns:
            {
                "table_name": str,
                "schema": dict,          # 表结构信息
                "row_count": int,
                "sample_data": list,     # 采样数据
                "duckdb_schema_sql": str, # CREATE TABLE 语句
            }
        """
        import duckdb

        df, meta = self.load_excel_smart(file_path, sheet_name=sheet_name)

        if df.empty:
            return {
                "table_name": table_name,
                "schema": {"columns": []},
                "row_count": 0,
                "sample_data": [],
                "duckdb_schema_sql": "",
                "error": "Excel 文件为空或无法解析",
            }

        conn = duckdb.connect(":memory:")
        conn.register("df_temp", df)
        conn.execute(f"CREATE TABLE {table_name} AS SELECT * FROM df_temp")
        conn.unregister("df_temp")

        # 提取 schema
        schema_result = conn.execute(f"DESCRIBE {table_name}").fetchall()
        columns_info = []
        for row in schema_result:
            columns_info.append({
                "name": row[0],
                "type": row[1],
            })

        # 采样数据
        sample = conn.execute(f"SELECT * FROM {table_name} LIMIT 5").fetchdf()
        sample_data = sample.to_dict(orient="records")

        row_count = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]

        # 生成 schema 描述（用于 LLM prompt）
        schema_lines = []
        for col in columns_info:
            schema_lines.append(f"  {col['name']} {col['type']}")
        schema_sql = f"CREATE TABLE {table_name} (\n" + ",\n".join(schema_lines) + "\n)"

        # 将连接存入类级缓存（按 table_name 索引）
        if not hasattr(self, "_duckdb_connections"):
            self._duckdb_connections = {}
        self._duckdb_connections[table_name] = conn

        logger.info(
            f"DuckDB: loaded {file_path} → table '{table_name}', "
            f"{row_count} rows, {len(columns_info)} columns"
        )

        return {
            "table_name": table_name,
            "schema": {"columns": columns_info},
            "row_count": row_count,
            "sample_data": sample_data,
            "duckdb_schema_sql": schema_sql,
            "load_metadata": meta,
        }

    def execute_duckdb_query(
        self, sql: str, table_name: str = "excel_data"
    ) -> Dict[str, Any]:
        """
        执行 DuckDB SQL 查询

        Returns:
            {
                "success": bool,
                "columns": list,
                "data": list,           # 查询结果（字典列表）
                "row_count": int,
                "error": Optional[str],
            }
        """
        conn = getattr(self, "_duckdb_connections", {}).get(table_name)
        if conn is None:
            return {
                "success": False,
                "columns": [],
                "data": [],
                "row_count": 0,
                "error": f"DuckDB table '{table_name}' not found. Please load data first.",
            }

        try:
            result_df = conn.execute(sql).fetchdf()
            columns = result_df.columns.tolist()
            data = result_df.to_dict(orient="records")

            # 安全转换非序列化类型
            for row in data:
                for key, val in row.items():
                    if hasattr(val, "isoformat"):
                        row[key] = val.isoformat()
                    elif pd.isna(val):
                        row[key] = None

            return {
                "success": True,
                "columns": columns,
                "data": data,
                "row_count": len(data),
                "error": None,
            }
        except Exception as e:
            logger.error(f"DuckDB query failed: {e}\nSQL: {sql}")
            return {
                "success": False,
                "columns": [],
                "data": [],
                "row_count": 0,
                "error": str(e),
            }

    def close_duckdb(self, table_name: str = "excel_data"):
        """关闭 DuckDB 连接，释放资源"""
        conn = getattr(self, "_duckdb_connections", {}).pop(table_name, None)
        if conn:
            conn.close()
            logger.info(f"DuckDB: closed connection for table '{table_name}'")

    def _extract_schema_sync(
        self,
        file_path: str,
        sheet_name: Optional[str],
        sample_rows: int,
    ) -> Dict[str, Any]:
        if file_path.lower().endswith('.csv'):
            sheet_names = ["Sheet1"]
            target = "Sheet1"
            df = self._load_dataframe(file_path)
        else:
            xl = pd.ExcelFile(file_path, engine="openpyxl")
            sheet_names = xl.sheet_names
            target = sheet_name or sheet_names[0]
            df = pd.read_excel(xl, sheet_name=target)

        columns = df.columns.tolist()
        dtypes = {col: str(dt) for col, dt in df.dtypes.items()}
        null_counts = df.isnull().sum().to_dict()
        unique_counts = {col: int(df[col].nunique()) for col in columns}
        sample = df.head(sample_rows)

        schema_json = {
            "file": os.path.basename(file_path),
            "sheet": target,
            "shape": list(df.shape),
            "columns": [],
        }
        for col in columns:
            schema_json["columns"].append({
                "name": col,
                "dtype": dtypes[col],
                "null_count": int(null_counts.get(col, 0)),
                "unique_count": unique_counts.get(col, 0),
                "sample_values": _safe_values(df[col].head(sample_rows)),
            })

        md_lines = [
            f"## File: {os.path.basename(file_path)} (Sheet: {target})",
            f"**Shape**: {df.shape[0]} rows × {df.shape[1]} columns",
            "",
            "| # | Column | Type | Nulls | Unique | Sample Values |",
            "|---|--------|------|-------|--------|---------------|",
        ]
        for i, col in enumerate(columns):
            samples = ", ".join(str(v) for v in _safe_values(df[col].head(3)))
            md_lines.append(
                f"| {i+1} | {col} | {dtypes[col]} | {null_counts.get(col, 0)} | "
                f"{unique_counts.get(col, 0)} | {samples} |"
            )

        md_lines.append("")
        md_lines.append("### Sample Data (first {} rows)".format(min(sample_rows, len(df))))
        md_lines.append(sample.to_markdown(index=False))

        stats = {}
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        if numeric_cols:
            desc = df[numeric_cols].describe()
            for col in numeric_cols:
                stats[col] = {
                    "mean": _safe_float(desc.loc["mean", col]) if "mean" in desc.index else None,
                    "std": _safe_float(desc.loc["std", col]) if "std" in desc.index else None,
                    "min": _safe_float(desc.loc["min", col]) if "min" in desc.index else None,
                    "max": _safe_float(desc.loc["max", col]) if "max" in desc.index else None,
                }

        return {
            "schema_markdown": "\n".join(md_lines),
            "schema_json": schema_json,
            "columns": columns,
            "dtypes": dtypes,
            "null_counts": null_counts,
            "unique_counts": unique_counts,
            "sample_data": sample.to_dict(orient="records"),
            "shape": df.shape,
            "sheet_names": sheet_names,
            "stats": stats,
        }

    def _execute_python_sync(self, file_path: str, code: str, user_instruction: str = "") -> Dict[str, Any]:
        try:
            df = self._load_dataframe(file_path)
        except Exception as e:
            return {"success": False, "error": f"Failed to read file: {e}", "output": "", "image_base64": None}

        # 强制注入常用数据分析和绘图库，防止 LLM 代码引用未定义变量 (NameError)
        import numpy as np

        # 配置 matplotlib 中文字体后备方案（Linux 通常缺少 SimHei，使用 Noto Sans CJK / Droid Sans Fallback）
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Noto Sans CJK JP', 'Droid Sans Fallback', 'DejaVu Sans', 'sans-serif']
        plt.rcParams['axes.unicode_minus'] = False

        sandbox_globals = {
            "__builtins__": __builtins__,
            "pd": pd,
            "plt": plt,
            "np": np,
            "df": df,
            "io": io,
            "base64": base64,
            "matplotlib": matplotlib,
        }

        output_capture = io.StringIO()

        try:
            with contextlib.redirect_stdout(output_capture):
                exec(code, sandbox_globals)

            output = output_capture.getvalue()

            # 优先提取 LLM 代码中显式赋值的 chart_base64 变量
            # 兼容 LLM 可能使用的不同变量名
            img_base64 = None
            for var_name in ("chart_base64", "img_base64", "chart_b64", "image_base64", "fig_base64", "base64_img"):
                val = sandbox_globals.get(var_name)
                if val and isinstance(val, str) and len(val) > 100:
                    img_base64 = val
                    logger.info(f"📊 [Sandbox] 图表从变量 '{var_name}' 提取成功 (len={len(val)})")
                    break

            # 如果没有 chart_base64，检查 matplotlib 是否有图形待保存
            if not img_base64 and plt.get_fignums():
                img_buf = io.BytesIO()
                plt.savefig(img_buf, format='png', bbox_inches='tight', dpi=150)
                img_buf.seek(0)
                img_base64 = base64.b64encode(img_buf.read()).decode('utf-8')
                logger.info(f"📊 [Sandbox] 图表从 plt.get_fignums() 自动捕获成功 (len={len(img_base64)})")

            if not img_base64:
                logger.warning("📊 [Sandbox] 未检测到图表输出（chart_base64 变量为空且 plt 无活动图形）")

            return {"success": True, "output": output, "image_base64": img_base64}

        except Exception:
            err_msg = traceback.format_exc()
            return {"success": False, "error": err_msg, "output": "", "image_base64": None}
        finally:
            # 物理销毁所有残留画布，防止内存泄漏与多次执行时图表重叠
            if 'plt' in sandbox_globals:
                try:
                    sandbox_globals['plt'].close('all')
                except Exception:
                    pass

    def _execute_pandas_sync(
        self,
        code: str,
        input_path: str,
        output_path: str,
    ) -> Dict[str, Any]:
        logs: List[str] = []

        def _print_capture(*args, **kwargs):
            logs.append(" ".join(str(a) for a in args))

        _code_check = _check_code_safety(code)
        if _code_check:
            return {
                "success": False,
                "output_path": None,
                "error": _code_check,
                "log": logs,
                "original_shape": None,
                "result_shape": None,
            }

        try:
            df = self._load_dataframe(input_path)
            original_shape = df.shape
            logger.info(f"Sandbox: loaded {input_path}, shape={original_shape}")
        except Exception as e:
            return {
                "success": False,
                "output_path": None,
                "error": f"Failed to read input file: {e}",
                "log": logs,
                "original_shape": None,
                "result_shape": None,
            }

        import numpy as np

        # 配置 matplotlib 中文字体后备方案（与 _execute_python_sync 保持一致）
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Noto Sans CJK JP', 'Droid Sans Fallback', 'DejaVu Sans', 'sans-serif']
        plt.rcParams['axes.unicode_minus'] = False

        sandbox_globals = {
            "__builtins__": __builtins__,
            **_SAFE_PANDAS,
            "np": np,
            "plt": plt,
            "pd": pd,
            "df": df,
            "io": io,
            "base64": base64,
            "matplotlib": matplotlib,
            "input_path": input_path,
            "output_path": output_path,
            "print": _print_capture,
        }

        try:
            exec(code, sandbox_globals)
        except Exception as e:
            tb = traceback.format_exc()
            logger.error(f"Sandbox exec failed: {e}\n{tb}")
            return {
                "success": False,
                "output_path": None,
                "error": f"{type(e).__name__}: {e}",
                "log": logs,
                "original_shape": original_shape,
                "result_shape": None,
            }

        try:
            result_df = sandbox_globals.get("df")
            if result_df is not None and isinstance(result_df, pd.DataFrame):
                try:
                    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
                    result_df.to_excel(output_path, index=False, engine="openpyxl")
                    logger.info(f"Sandbox: saved result to {output_path}, shape={result_df.shape}")
                except Exception as e:
                    return {
                        "success": False,
                        "output_path": None,
                        "error": f"Failed to save output: {e}",
                        "log": logs,
                        "original_shape": original_shape,
                        "result_shape": list(result_df.shape) if isinstance(result_df, pd.DataFrame) else None,
                    }
            else:
                if os.path.exists(output_path):
                    logger.info(f"Sandbox: code saved output via output_path variable")
                else:
                    return {
                        "success": False,
                        "output_path": None,
                        "error": "Code did not produce a valid DataFrame (df) or save to output_path",
                        "log": logs,
                        "original_shape": original_shape,
                        "result_shape": None,
                    }

            result_shape = None
            if isinstance(result_df, pd.DataFrame):
                result_shape = list(result_df.shape)

            return {
                "success": True,
                "output_path": output_path,
                "error": None,
                "log": logs,
                "original_shape": original_shape,
                "result_shape": result_shape,
            }
        finally:
            # 物理销毁所有残留画布，防止内存泄漏与多次执行时图表重叠
            if 'plt' in sandbox_globals:
                try:
                    sandbox_globals['plt'].close('all')
                except Exception:
                    pass

    def _load_sync(
        self, file_path: str, sheet_name: Optional[str], header: int
    ) -> Dict[str, Any]:
        if file_path.lower().endswith('.csv'):
            sheet_names = ["Sheet1"]
            df = self._load_dataframe(file_path)
        else:
            xl = pd.ExcelFile(file_path, engine="openpyxl")
            sheet_names = xl.sheet_names
            target = sheet_name or sheet_names[0]
            df = pd.read_excel(xl, sheet_name=target, header=header)
        preview = df.head(20).to_dict(orient="records")
        return {
            "dataframe": df,
            "shape": df.shape,
            "columns": df.columns.tolist(),
            "dtypes": {col: str(dt) for col, dt in df.dtypes.items()},
            "sheet_names": sheet_names,
            "null_counts": df.isnull().sum().to_dict(),
            "preview": preview,
        }

    def _get_sheets_sync(self, file_path: str) -> List[str]:
        if file_path.lower().endswith('.csv'):
            return ["Sheet1"]
        xl = pd.ExcelFile(file_path, engine="openpyxl")
        return xl.sheet_names


_SAFE_PANDAS = {
    "pd": pd,
    "DataFrame": pd.DataFrame,
    "Series": pd.Series,
    "NA": pd.NA,
    "Timestamp": pd.Timestamp,
    "Timedelta": pd.Timedelta,
    "to_datetime": pd.to_datetime,
    "to_numeric": pd.to_numeric,
    "concat": pd.concat,
    "merge": pd.merge,
    "melt": pd.melt,
    "pivot_table": pd.pivot_table,
    "cut": pd.cut,
    "qcut": pd.qcut,
}


def _safe_values(series: pd.Series) -> list:
    result = []
    for v in series:
        if pd.isna(v):
            result.append("NaN")
        else:
            result.append(str(v))
    return result


def _safe_float(val) -> Optional[float]:
    try:
        if pd.isna(val):
            return None
        return round(float(val), 4)
    except (ValueError, TypeError):
        return None


def _check_code_safety(code: str) -> Optional[str]:
    stripped_lines = []
    for line in code.split("\n"):
        s = line.strip()
        if s and not s.startswith("#"):
            stripped_lines.append(s)

    for line in stripped_lines:
        tokens = line.split()
        if not tokens:
            continue

        first = tokens[0]
        if first == "import" and len(tokens) >= 2:
            module = tokens[1].split(".")[0]
            if module in _FORBIDDEN_MODULES:
                return f"Forbidden import: {module}"
        if first == "from" and len(tokens) >= 2:
            module = tokens[1].split(".")[0]
            if module in _FORBIDDEN_MODULES:
                return f"Forbidden import from: {module}"

        for name in _FORBIDDEN_NAMES:
            if f" {name}(" in line or line.startswith(f"{name}("):
                return f"Forbidden function: {name}"

    return None
